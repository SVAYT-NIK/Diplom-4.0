"""
Excel file parser for heat consumption data.
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import load_workbook

from backend.models.schemas import DailyReadingBase


logger = logging.getLogger(__name__)


def normalize_date(date_value: Any) -> Optional[str]:
    """
    Normalize date values to YYYY-MM-DD format.
    
    Args:
        date_value: Date value in various formats (DD.MM.YYYY, MM/DD/YY, datetime, etc.)
    
    Returns:
        Normalized date string in YYYY-MM-DD format or None if parsing fails
    """
    if pd.isna(date_value) or date_value is None:
        return None
    
    # If it's already a datetime object
    if isinstance(date_value, datetime):
        return date_value.strftime("%Y-%m-%d")
    
    # If it's a pandas Timestamp
    if isinstance(date_value, pd.Timestamp):
        return date_value.strftime("%Y-%m-%d")
    
    # Convert to string for parsing
    date_str = str(date_value).strip()
    
    if not date_str:
        return None
    
    # Try different date formats
    date_formats = [
        "%d.%m.%Y",  # DD.MM.YYYY
        "%d.%m.%y",  # DD.MM.YY
        "%m/%d/%Y",  # MM/DD/YYYY
        "%m/%d/%y",  # MM/DD/YY
        "%Y-%m-%d",  # YYYY-MM-DD (already normalized)
        "%d-%m-%Y",  # DD-MM-YYYY
    ]
    
    for fmt in date_formats:
        try:
            parsed_date = datetime.strptime(date_str, fmt)
            return parsed_date.strftime("%Y-%m-%d")
        except ValueError:
            continue
    
    # If all formats fail, log warning and return None
    logger.warning(f"Could not parse date: {date_str}")
    return None


def parse_ns_codes(ns_value: Any) -> Optional[str]:
    """
    Parse NS (non-standard situation) codes from various formats.
    
    Args:
        ns_value: NS codes value (can be comma-separated, single value, etc.)
    
    Returns:
        Comma-separated string of NS codes or None
    """
    if pd.isna(ns_value) or ns_value is None:
        return None
    
    ns_str = str(ns_value).strip()
    
    if not ns_str or ns_str.lower() in ['nan', 'none', 'null', '']:
        return None
    
    # If it contains commas, split and rejoin to clean up
    if ',' in ns_str:
        codes = [code.strip() for code in ns_str.split(',')]
        codes = [code for code in codes if code]  # Remove empty strings
        return ','.join(codes) if codes else None
    
    return ns_str


def clean_numeric(value: Any) -> Optional[float]:
    """
    Clean numeric values by replacing commas with dots and handling NaN.
    
    Args:
        value: Numeric value that may contain commas or be NaN
    
    Returns:
        Float value or None if invalid
    """
    if pd.isna(value) or value is None:
        return None
    
    # Convert to string first
    value_str = str(value).strip()
    
    if not value_str or value_str.lower() in ['nan', 'none', 'null', '']:
        return None
    
    # Replace comma with dot for decimal separator
    value_str = value_str.replace(',', '.')
    
    try:
        return float(value_str)
    except (ValueError, TypeError):
        logger.warning(f"Could not convert to float: {value_str}")
        return None


def parse_excel_file(file_path: Path) -> tuple[Optional[Dict[str, Any]], List[DailyReadingBase]]:
    """
    Parse Excel file with heat consumption data.
    
    Expected format:
    - Rows 1-4: Metadata (report info, calculator, consumer, scheme)
    - Row 5: Headers
    - Row 6+: Data
    - Special rows to ignore: "Итого за период штатной работы", "Итого за период НС", 
      "Итого", "Нештатные ситуации"
    
    Args:
        file_path: Path to Excel file
    
    Returns:
        Tuple of (metadata dict, list of DailyReadingBase objects)
    """
    logger.info(f"Parsing Excel file: {file_path}")
    
    try:
        # Load workbook to extract metadata
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        
        # Extract metadata from first 4 rows
        metadata = {
            "report_generated": None,
            "calculator": None,
            "consumer": None,
            "scheme": None,
        }
        
        # Try to extract metadata (adjust based on actual file structure)
        try:
            if ws.cell(row=1, column=1).value:
                metadata["report_generated"] = str(ws.cell(row=1, column=1).value)
            if ws.cell(row=2, column=1).value:
                metadata["calculator"] = str(ws.cell(row=2, column=1).value)
            if ws.cell(row=3, column=1).value:
                metadata["consumer"] = str(ws.cell(row=3, column=1).value)
            if ws.cell(row=4, column=1).value:
                metadata["scheme"] = str(ws.cell(row=4, column=1).value)
        except Exception as e:
            logger.warning(f"Could not extract metadata: {e}")
        
        wb.close()
        
        # Read data with pandas starting from row 6 (0-indexed: row 5)
        # First read all data without skipping to get headers correctly
        df_raw = pd.read_excel(
            file_path,
            sheet_name=0,
            header=None,  # Don't use any row as header initially
        )
        
        # Extract headers from row 5 (0-indexed: row 4)
        if len(df_raw) > 4:
            headers = df_raw.iloc[4].tolist()
            # Create new dataframe with data starting from row 6 (0-indexed: row 5)
            df = pd.DataFrame(columns=headers)
            if len(df_raw) > 5:
                data_rows = df_raw.iloc[5:].copy()
                data_rows.columns = headers
                df = data_rows.reset_index(drop=True)
        else:
            logger.error("Excel file doesn't have enough rows for headers")
            raise ValueError("Excel file must have at least 5 rows (4 metadata + 1 header)")
        
        logger.info(f"Loaded DataFrame with shape: {df.shape}")
        logger.debug(f"Columns: {df.columns.tolist()}")
        
        # Filter out summary rows
        summary_keywords = [
            "Итого за период штатной работы",
            "Итого за период НС",
            "Итого",
            "Нештатные ситуации",
        ]
        
        # Check if 'Дата' column exists and filter
        if 'Дата' in df.columns:
            for keyword in summary_keywords:
                df = df[~df['Дата'].astype(str).str.contains(keyword, na=False)]
        
        readings = []
        
        for idx, row in df.iterrows():
            try:
                # Skip rows with invalid dates
                date_normalized = normalize_date(row.get('Дата'))
                if not date_normalized:
                    continue
                
                # Create reading object
                reading = DailyReadingBase(
                    date=date_normalized,
                    t1=clean_numeric(row.get('T1')),
                    t2=clean_numeric(row.get('T2')),
                    p1=clean_numeric(row.get('P1')),
                    p2=clean_numeric(row.get('P2')),
                    v1=clean_numeric(row.get('V1')),
                    v2=clean_numeric(row.get('V2')),
                    m1=clean_numeric(row.get('M1')),
                    m2=clean_numeric(row.get('M2')),
                    q=clean_numeric(row.get('Q')),
                    dt=clean_numeric(row.get('d T')),
                    dv=clean_numeric(row.get('d V')),
                    dm=clean_numeric(row.get('d M')),
                    imbalance=clean_numeric(row.get('Небаланс')),
                    ns_codes=parse_ns_codes(row.get('НС')),
                    status=str(row.get('Состояние')) if pd.notna(row.get('Состояние')) else None,
                )
                
                readings.append(reading)
                
            except Exception as e:
                logger.warning(f"Error parsing row {idx}: {e}")
                continue
        
        logger.info(f"Successfully parsed {len(readings)} readings")
        
        return metadata, readings
        
    except Exception as e:
        logger.error(f"Failed to parse Excel file: {e}")
        raise


def get_building_info_from_metadata(metadata: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract building information from metadata.
    
    Args:
        metadata: Dictionary with metadata from Excel file
    
    Returns:
        Dictionary with building information
    """
    building_info = {
        "address": None,
        "area_m2": None,
        "year_built": None,
        "heating_type": "central",
    }
    
    # Try to extract consumer info (usually contains address)
    if metadata.get("consumer"):
        consumer_str = str(metadata["consumer"])
        # Simple heuristic: first part might be address
        building_info["address"] = consumer_str.split(',')[0].strip()
    
    # Try to extract area from scheme or other fields
    if metadata.get("scheme"):
        scheme_str = str(metadata["scheme"])
        # Look for patterns like "площадь 1234 м2" or "S=1234"
        import re
        area_match = re.search(r'(?:площадь|S\s*=?\s*)(\d+(?:[.,]\d+)?)', scheme_str, re.IGNORECASE)
        if area_match:
            area_str = area_match.group(1).replace(',', '.')
            building_info["area_m2"] = float(area_str)
    
    return building_info
